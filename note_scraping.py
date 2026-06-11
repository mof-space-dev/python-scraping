from selenium import webdriver
import pandas as pd
from bs4 import BeautifulSoup
import time

import openpyxl
from openpyxl.styles import PatternFill, Font

driver = webdriver.Chrome()
driver.get("https://note.com/hashtag/python")
time.sleep(3)

data = []

before = 0 # スクロール前の件数
while True:
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(2)
    
    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")
    after = soup.find_all("a", title=True)
    after = len(after)
    
    if before == after: # after スクロール後の件数
       break
    
    before = after

# ループが終わってから全件取得
items = soup.find_all("a", title=True)
for item in items:
    data.append({
      "タイトル": item['title'],
      "URL": f"https://note.com{item['href']}"
    })

df = pd.DataFrame(data)
df.to_excel("note_scraping_result.xlsx", index=False)
driver.quit()
print(f"{len(df)}件保存完了")

# # ヘッダー色付け（青・白文字・太字）
# 列幅は自動調整
# 縞模様
# 1行目固定

wb = openpyxl.load_workbook("note_scraping_result.xlsx")
ws = wb.active

# ヘッダー
blue = PatternFill(start_color="0000ff", end_color="0000ff", fill_type="solid")
white = Font(color="ffffff", bold=True)
for cell in ws[1]:
    cell.fill = blue
    cell.font = white

# 列幅調整
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
# 縞模様
stripe = PatternFill(start_color="fff0f5", end_color="fff0f5", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[0].row % 2 == 1:
      for cell in row:
          cell.fill = stripe
          
# ヘッダー固定
ws.freeze_panes = "A2"
wb.save("note_scraping_result.xlsx")